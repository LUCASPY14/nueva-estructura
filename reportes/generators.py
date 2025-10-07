from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional

from django.db.models import Sum, Count, Avg, F, Q, QuerySet, Prefetch
from django.utils import timezone
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import json
import csv
from io import BytesIO, StringIO

from ventas.models import Venta, DetalleVenta
from productos.models import Producto, Categoria
from alumnos.models import Alumno
from .cache import ReportCache

class BaseReportGenerator(ABC):
    """
    Clase base abstracta para generadores de reportes.
    Define la interfaz común para todos los generadores de reportes.
    """
    
    def __init__(self, fecha_inicio: datetime, fecha_fin: datetime, **kwargs):
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        self.extra_params = kwargs
        
    @abstractmethod
    def generate_data(self) -> Dict[str, Any]:
        """Genera los datos del reporte"""
        pass
        
    @abstractmethod
    def get_template_name(self) -> str:
        """Retorna el nombre del template para el reporte PDF"""
        pass
    
    def generate_pdf(self, context: Dict[str, Any]) -> bytes:
        """Genera un reporte en formato PDF"""
        from django.template.loader import render_to_string
        
        html_string = render_to_string(self.get_template_name(), context)
        html = HTML(string=html_string)
        return html.write_pdf()
    
    def generate_excel(self, data: Dict[str, Any]) -> bytes:
        """Genera un reporte en formato Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Implementación específica para cada tipo de reporte
        self._write_excel_data(ws, data)
        
        # Guardar
        excel_file = BytesIO()
        wb.save(excel_file)
        return excel_file.getvalue()
    
    @abstractmethod
    def _write_excel_data(self, worksheet: Any, data: Dict[str, Any]) -> None:
        """Implementación específica para escribir datos en Excel"""
        pass

class VentasReportGenerator(BaseReportGenerator):
    """Generador de reportes de ventas"""
    
    def generate_data(self) -> Dict[str, Any]:
        ventas = Venta.objects.filter(
            fecha__range=(self.fecha_inicio, self.fecha_fin),
            estado='completada'
        ).select_related('alumno', 'cajero')
        
        # Resumen general
        resumen = ventas.aggregate(
            total_ventas=Count('id'),
            monto_total=Sum('total'),
            ticket_promedio=Avg('total')
        )
        
        # Ventas por día
        ventas_por_dia = ventas.values('fecha__date').annotate(
            total=Sum('total'),
            cantidad=Count('id')
        ).order_by('fecha__date')
        
        # Productos más vendidos
        productos_vendidos = DetalleVenta.objects.filter(
            venta__in=ventas
        ).values('producto__nombre').annotate(
            cantidad=Sum('cantidad'),
            total=Sum('subtotal')
        ).order_by('-cantidad')[:10]
        
        return {
            'resumen': resumen,
            'ventas_por_dia': list(ventas_por_dia),
            'productos_vendidos': list(productos_vendidos),
            'fecha_inicio': self.fecha_inicio,
            'fecha_fin': self.fecha_fin
        }
    
    def get_template_name(self) -> str:
        return 'reportes/pdf/reporte_ventas.html'
    
    def _write_excel_data(self, ws: Any, data: Dict[str, Any]) -> None:
        # Encabezado del reporte
        ws['A1'] = "Reporte de Ventas"
        ws['A2'] = f"Período: {data['fecha_inicio'].strftime('%d/%m/%Y')} - {data['fecha_fin'].strftime('%d/%m/%Y')}"
        
        # Resumen
        ws['A4'] = "Resumen"
        headers = [
            ['Total Ventas', data['resumen']['total_ventas']],
            ['Monto Total', f"${data['resumen']['monto_total']:.2f}"],
            ['Ticket Promedio', f"${data['resumen']['ticket_promedio']:.2f}"]
        ]
        for i, (label, value) in enumerate(headers, 5):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            
        # Ventas por día
        ws['A7'] = "Ventas por Día"
        ws['A8'] = "Fecha"
        ws['B8'] = "Cantidad"
        ws['C8'] = "Total"
        
        for i, venta in enumerate(data['ventas_por_dia'], 9):
            ws[f'A{i}'] = venta['fecha__date'].strftime('%d/%m/%Y')
            ws[f'B{i}'] = venta['cantidad']
            ws[f'C{i}'] = f"${venta['total']:.2f}"

class InventarioReportGenerator(BaseReportGenerator):
    """Generador de reportes de inventario"""
    
    def generate_data(self) -> Dict[str, Any]:
        productos = Producto.objects.all().select_related('categoria')
        
        # Resumen general
        resumen = productos.aggregate(
            total_productos=Count('id'),
            valor_total=Sum(F('cantidad') * F('precio_costo')),
            productos_sin_stock=Count('id', filter=Q(cantidad=0)),
            productos_bajo_stock=Count('id', filter=Q(cantidad__lte=F('cantidad_minima')))
        )
        
        # Productos por categoría
        por_categoria = productos.values('categoria__nombre').annotate(
            cantidad=Count('id'),
            valor=Sum(F('cantidad') * F('precio_costo'))
        ).order_by('-cantidad')
        
        # Productos críticos
        criticos = productos.filter(
            Q(cantidad=0) | Q(cantidad__lte=F('cantidad_minima'))
        ).values('nombre', 'codigo', 'cantidad', 'cantidad_minima')
        
        return {
            'resumen': resumen,
            'por_categoria': list(por_categoria),
            'criticos': list(criticos),
            'fecha_reporte': timezone.now()
        }
    
    def get_template_name(self) -> str:
        return 'reportes/pdf/reporte_inventario.html'
    
    def _write_excel_data(self, ws: Any, data: Dict[str, Any]) -> None:
        # Similar a VentasReportGenerator pero con datos de inventario
        pass